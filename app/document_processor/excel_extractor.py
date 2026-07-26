import openpyxl

def extract_text_from_excel(filepath: str) -> str:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    text_parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        text_parts.append(f"\n--- Hoja: {sheet_name} ---\n")
        rows_text = []
        for row in ws.iter_rows(values_only=True):
            row_str = " | ".join(str(cell) if cell is not None else "" for cell in row)
            if row_str.strip() and row_str.strip() != " | " * (len(row) - 1):
                rows_text.append(row_str)
        text_parts.append("\n".join(rows_text))
    return "\n".join(text_parts)
